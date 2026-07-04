import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

def no_smooth(chart):
    for s in chart.series:
        s.smooth = False

wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

# =========================================================
# HOJA P1 - Carta np (discos fallidos por rack)
# =========================================================
ws = wb.create_sheet("P1 Carta np")
ws["A1"] = "PROBLEMA 1 — Discos duros fallidos por rack (carta np, n=240 constante)"
ws["A1"].font = TITLE_FONT

racks = list(range(1, 21))
fallidos = [3,5,2,4,6,3,5,4,7,3,4,5,2,6,4,5,3,4,6,5]

ws["A3"] = "Rack"
ws["B3"] = "Discos fallidos"
style_header_row(ws, 3, 1, 2)

for i, (r, f) in enumerate(zip(racks, fallidos)):
    row = 4 + i
    ws.cell(row=row, column=1, value=r).border = BORDER
    ws.cell(row=row, column=2, value=f).border = BORDER

last_data_row = 3 + len(racks)  # 23

# Parametros
ws["D3"] = "Parámetro"
ws["E3"] = "Valor"
ws["F3"] = "Fórmula"
style_header_row(ws, 3, 4, 6)

ws["D4"] = "n (tamaño de muestra por rack)"
ws["E4"] = 240
ws["D5"] = "k (número de racks)"
ws["E5"] = f"=COUNT(A4:A{last_data_row})"
ws["D6"] = "np̄ (promedio de defectuosos)"
ws["E6"] = f"=AVERAGE(B4:B{last_data_row})"
ws["F6"] = f"=AVERAGE(B4:B{last_data_row})"
ws["D7"] = "p̄ = np̄ / n"
ws["E7"] = "=E6/E4"
ws["D8"] = "UCL = np̄ + 3·SQRT(np̄·(1-p̄))"
ws["E8"] = "=E6+3*SQRT(E6*(1-E7))"
ws["D9"] = "LCL = MAX(0, np̄ - 3·SQRT(np̄·(1-p̄)))"
ws["E9"] = "=MAX(0,E6-3*SQRT(E6*(1-E7)))"
ws["D10"] = "CL = np̄"
ws["E10"] = "=E6"

for row in range(4, 11):
    ws.cell(row=row, column=4).font = LABEL_FONT

# Columns for chart: UCL/CL/LCL constants replicated
ws["H3"] = "Rack"
ws["I3"] = "Fallidos"
ws["J3"] = "UCL"
ws["K3"] = "CL(np̄)"
ws["L3"] = "LCL"
style_header_row(ws, 3, 8, 12)
for i, r in enumerate(racks):
    row = 4 + i
    ws.cell(row=row, column=8, value=r).border = BORDER
    ws.cell(row=row, column=9, value=f"=B{row}").border = BORDER
    ws.cell(row=row, column=10, value="=$E$8").border = BORDER
    ws.cell(row=row, column=11, value="=$E$10").border = BORDER
    ws.cell(row=row, column=12, value="=$E$9").border = BORDER

chart = LineChart()
chart.title = "Carta np — Discos fallidos por rack"
chart.y_axis.title = "Discos fallidos"
chart.x_axis.title = "Rack"
data = Reference(ws, min_col=9, max_col=12, min_row=3, max_row=last_data_row)
cats = Reference(ws, min_col=8, min_row=4, max_row=last_data_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 9
chart.width = 20
no_smooth(chart)
ws.add_chart(chart, "D13")

for col, w in zip("ABCDEFGHIJKL", [8,16,8,32,32,10,4,8,10,10,10,10]):
    ws.column_dimensions[col].width = w

# =========================================================
# HOJA P2 - Capacidad Taguchi
# =========================================================
ws = wb.create_sheet("P2 Capacidad")
ws["A1"] = "PROBLEMA 2 — Tiempo de respuesta de servidor (capacidad + índice Taguchi)"
ws["A1"].font = TITLE_FONT

ws["A3"] = "Estadístico"
ws["B3"] = "Valor"
style_header_row(ws, 3, 1, 2)

inputs = [
    ("Media (X̄)", 105),
    ("Desv. estándar (s)", 18),
    ("n", 180),
    ("Target (T)", 100),
    ("LSL", 50),
    ("USL", 150),
]
for i, (label, val) in enumerate(inputs):
    row = 4 + i
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    ws.cell(row=row, column=2, value=val).border = BORDER

# named-ish references
XBAR, S, N, T, LSL, USL = "B4", "B5", "B6", "B7", "B8", "B9"

ws["A11"] = "Índice"
ws["B11"] = "Fórmula Excel"
ws["C11"] = "Resultado"
style_header_row(ws, 11, 1, 3)

rows_formulas = [
    ("Cp",  f"=(({USL})-({LSL}))/(6*{S})"),
    ("Cpl (Cpi)", f"=(({XBAR})-({LSL}))/(3*{S})"),
    ("Cpu (Cps)", f"=(({USL})-({XBAR}))/(3*{S})"),
    ("Cpk", f"=MIN(B13,B14)"),
    ("Cpm (Taguchi)", f"=(({USL})-({LSL}))/(6*SQRT({S}^2+(({XBAR})-({T}))^2))"),
]
for i, (name, formula) in enumerate(rows_formulas):
    row = 12 + i
    ws.cell(row=row, column=1, value=name).font = LABEL_FONT
    ws.cell(row=row, column=2, value=formula).border = BORDER
    ws.cell(row=row, column=3, value=formula.replace("B13","B13").replace("B14","B14"))  # placeholder, fixed below

# fix: put actual formulas in column C (resultado), column B as text description swapped
# Simplify: column B = formula text (as label), column C = live formula
ws["A11"] = "Índice"
ws["B11"] = "Fórmula"
ws["C11"] = "Resultado"

defs = [
    ("Cp",  "(USL-LSL)/(6·s)", f"=(({USL})-({LSL}))/(6*{S})"),
    ("Cpl (Cpi)", "(X̄-LSL)/(3·s)", f"=(({XBAR})-({LSL}))/(3*{S})"),
    ("Cpu (Cps)", "(USL-X̄)/(3·s)", f"=(({USL})-({XBAR}))/(3*{S})"),
    ("Cpk", "MIN(Cpu,Cpl)", None),  # filled after row numbers known
    ("Cpm (Taguchi)", "(USL-LSL)/(6·SQRT(s²+(X̄-T)²))", f"=(({USL})-({LSL}))/(6*SQRT({S}^2+(({XBAR})-({T}))^2))"),
]
start_row = 12
for i, (name, desc, formula) in enumerate(defs):
    row = start_row + i
    ws.cell(row=row, column=1, value=name).font = LABEL_FONT
    ws.cell(row=row, column=2, value=desc)
    if formula is not None:
        ws.cell(row=row, column=3, value=formula)
    ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=2).border = BORDER
    ws.cell(row=row, column=3).border = BORDER

cpu_row = start_row + 2  # Cpu row
cpl_row = start_row + 1  # Cpl row
cpk_row = start_row + 3
ws.cell(row=cpk_row, column=3, value=f"=MIN(C{cpu_row},C{cpl_row})")

ws["A19"] = "Diagnóstico rápido"
ws["A19"].font = LABEL_FONT
ws["A20"] = "Cp vs Cpk"
ws["B20"] = f"=IF(ABS(C{start_row}-C{cpk_row})<0.05,\"Proceso centrado\",\"Proceso descentrado\")"
ws["A21"] = "Nivel de capacidad (Cpk)"
ws["B21"] = f"=IF(C{cpk_row}<1,\"No capaz\",IF(C{cpk_row}<1.33,\"Capaz marginal\",\"Capaz\"))"

for col, w in zip("ABC", [22, 42, 14]):
    ws.column_dimensions[col].width = w

# Gráfico de barras comparando índices (categóricos, no serie de tiempo)
bar = BarChart()
bar.type = "col"
bar.title = "Índices de capacidad"
cats2 = Reference(ws, min_col=1, min_row=start_row, max_row=cpk_row+1)
data2 = Reference(ws, min_col=3, min_row=11, max_row=cpk_row+1)
bar.add_data(data2, titles_from_data=True)
bar.set_categories(cats2)
bar.height = 8
bar.width = 16
ws.add_chart(bar, "E12")

# =========================================================
# HOJA P3 - Carta X-R
# =========================================================
ws = wb.create_sheet("P3 Carta X-R")
ws["A1"] = "PROBLEMA 3 — Temperatura pasillo frío (carta X̄-R, n=4)"
ws["A1"].font = TITLE_FONT

data_p3 = [
    (21.8,22.1,21.9,22.0),
    (22.0,21.9,22.1,22.0),
    (21.9,22.2,22.0,21.8),
    (22.1,22.0,21.9,22.2),
    (22.0,22.3,22.1,21.9),
    (22.2,22.4,22.1,22.3),
    (22.3,22.1,22.4,22.2),
    (22.5,22.7,22.4,22.6),
    (22.4,22.6,22.3,22.5),
    (22.7,22.9,22.6,22.8),
    (22.6,22.4,22.7,22.5),
    (22.8,23.0,22.7,22.9),
    (22.9,23.1,22.8,23.0),
    (23.0,23.2,22.9,23.1),
    (23.1,23.3,23.0,23.2),
]

headers = ["Subgrupo","X1","X2","X3","X4","X̄","R"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=3, column=c, value=h)
style_header_row(ws, 3, 1, 7)

for i, vals in enumerate(data_p3):
    row = 4 + i
    ws.cell(row=row, column=1, value=i+1).border = BORDER
    for j, v in enumerate(vals):
        ws.cell(row=row, column=2+j, value=v).border = BORDER
    ws.cell(row=row, column=6, value=f"=AVERAGE(B{row}:E{row})").border = BORDER
    ws.cell(row=row, column=7, value=f"=MAX(B{row}:E{row})-MIN(B{row}:E{row})").border = BORDER

last_p3 = 3 + len(data_p3)  # 18

ws["I3"] = "Parámetro"
ws["J3"] = "Valor"
style_header_row(ws, 3, 9, 10)
ws["I4"] = "n (tamaño subgrupo)"
ws["J4"] = 4
ws["I5"] = "A2 (n=4)"
ws["J5"] = 0.729
ws["I6"] = "D3 (n=4)"
ws["J6"] = 0
ws["I7"] = "D4 (n=4)"
ws["J7"] = 2.282
ws["I8"] = "X̿ (promedio de X̄)"
ws["J8"] = f"=AVERAGE(F4:F{last_p3})"
ws["I9"] = "R̄ (promedio de R)"
ws["J9"] = f"=AVERAGE(G4:G{last_p3})"
ws["I10"] = "UCL X̄ = X̿ + A2·R̄"
ws["J10"] = "=J8+J5*J9"
ws["I11"] = "LCL X̄ = X̿ - A2·R̄"
ws["J11"] = "=J8-J5*J9"
ws["I12"] = "CL X̄ = X̿"
ws["J12"] = "=J8"
ws["I13"] = "UCL R = D4·R̄"
ws["J13"] = "=J7*J9"
ws["I14"] = "LCL R = D3·R̄"
ws["J14"] = "=J6*J9"
ws["I15"] = "CL R = R̄"
ws["J15"] = "=J9"
for row in range(4, 16):
    ws.cell(row=row, column=9).font = LABEL_FONT

# columnas de apoyo para graficar X-bar chart con límites constantes
ws["L3"] = "Subgrupo"
ws["M3"] = "X̄"
ws["N3"] = "UCL"
ws["O3"] = "CL"
ws["P3"] = "LCL"
style_header_row(ws, 3, 12, 16)
for i in range(len(data_p3)):
    row = 4 + i
    ws.cell(row=row, column=12, value=f"=A{row}").border = BORDER
    ws.cell(row=row, column=13, value=f"=F{row}").border = BORDER
    ws.cell(row=row, column=14, value="=$J$10").border = BORDER
    ws.cell(row=row, column=15, value="=$J$12").border = BORDER
    ws.cell(row=row, column=16, value="=$J$11").border = BORDER

chart3 = LineChart()
chart3.title = "Carta X̄ — Temperatura pasillo frío"
chart3.y_axis.title = "°C"
chart3.x_axis.title = "Subgrupo"
data3 = Reference(ws, min_col=13, max_col=16, min_row=3, max_row=last_p3)
cats3 = Reference(ws, min_col=12, min_row=4, max_row=last_p3)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.height = 9
chart3.width = 20
no_smooth(chart3)
ws.add_chart(chart3, "I17")

for col, w in zip("ABCDEFGHIJKLMNOP", [10,7,7,7,7,9,7,3,26,10,2,10,8,8,8,8]):
    ws.column_dimensions[col].width = w

# =========================================================
# HOJA P4 - Análisis crítico (conceptual)
# =========================================================
ws = wb.create_sheet("P4 Analisis Critico")
ws["A1"] = "PROBLEMA 4 — Análisis conceptual (sin datos de carta a graficar)"
ws["A1"].font = TITLE_FONT

ws["A3"] = "Sub-pregunta"
ws["B3"] = "Respuesta breve"
style_header_row(ws, 3, 1, 2)

ws["A4"] = "1. ¿Cpk alto pero fuera de control?"
ws["B4"] = ("Sí. Cpk usa la variación 'within-subgroup' (corto plazo) contra los límites de "
            "especificación; no verifica si el proceso está estadísticamente estable. Un servidor puede "
            "tener toda su variación dentro de LSL/USL (Cpk alto) y aun así mostrar una tendencia, "
            "tener un punto fuera de los límites de control, o rachas — es decir, causas especiales activas "
            "que Cpk no detecta porque solo compara dispersión vs. spec, no estabilidad en el tiempo.")

ws["A5"] = "2. Corto plazo (Cp/Cpk) vs largo plazo (Pp/Ppk)"
ws["B5"] = ("Cp/Cpk se calculan con la variación dentro de subgrupo (within), asumiendo que el proceso está "
            "en control estadístico en el momento del estudio — capacidad potencial de corto plazo. "
            "Pp/Ppk usan la desviación estándar global (overall) de todos los datos, capturando también "
            "la variación entre subgrupos (deriva, cambios de turno, degradación gradual). Para un SLA de "
            "uptime frente a clientes empresariales conviene reportar Pp/Ppk porque el cliente experimenta "
            "la variación real acumulada a lo largo de meses, no la variación idealizada de un subgrupo aislado.")

ws["A6"] = "3. Incidentes de seguridad por trimestre: ¿c o I-MR?"
ws["B6"] = ("Es un conteo de eventos raros (defectos) sin un denominador de 'éxito/fracaso' de tamaño fijo "
            "(no hay una base de 'n incidentes posibles por trimestre'), por lo que corresponde una carta c, "
            "no I-MR. I-MR asume una variable continua medida individualmente; aquí el dato es un conteo "
            "discreto de Poisson. Carta c: c̄=SUMA(incidentes)/k; UCL=c̄+3·RAIZ(c̄); LCL=MAX(0,c̄-3·RAIZ(c̄)).")

ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 100
for row in range(4, 7):
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 90

wb.save("solucion.xlsx")
print("OK: solucion.xlsx generado")
